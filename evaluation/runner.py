from pathlib import Path
import json
from statistics import mean

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from evaluation.adapters import (
    AgentAdapter,
    RAGAdapter,
)
from evaluation.dataset import load_dataset
from evaluation.metrics import (
    GeminiJudge,
    agent_metrics,
    retrieval_metrics,
    token_f1,
)


def _mean_numeric_metrics(
    results: list[dict],
) -> dict:
    """
    Calculate the mean of every numeric metric across
    successfully evaluated examples.
    """

    values = {}

    for result in results:
        for key, value in result.items():
            if isinstance(
                value,
                (int, float),
            ):
                values.setdefault(
                    key,
                    [],
                ).append(
                    float(value)
                )

    return {
        key: mean(items)
        for key, items in values.items()
        if items
    }


def _json_string(
    value,
) -> str:
    """
    Convert nested lists and dictionaries into JSON strings
    suitable for Excel cells.
    """

    return json.dumps(
        value,
        ensure_ascii=False,
        default=str,
    )


def _autosize_columns(
    worksheet,
    max_width: int = 50,
) -> None:
    """
    Automatically size worksheet columns while preventing
    very long text columns from becoming excessively wide.
    """

    for column_cells in worksheet.columns:
        max_length = 0

        column_letter = get_column_letter(
            column_cells[0].column
        )

        for cell in column_cells:
            value = cell.value

            if value is None:
                continue

            max_length = max(
                max_length,
                len(str(value)),
            )

        worksheet.column_dimensions[
            column_letter
        ].width = min(
            max_length + 2,
            max_width,
        )


def _style_header(
    worksheet,
) -> None:
    """
    Style the first row of a worksheet.
    """

    header_fill = PatternFill(
        fill_type="solid",
        fgColor="1F4E78",
    )

    for cell in worksheet[1]:
        cell.font = Font(
            bold=True,
            color="FFFFFF",
        )

        cell.fill = header_fill

        cell.alignment = Alignment(
            vertical="top",
            wrap_text=True,
        )


def _get_metric_by_prefix(
    metrics: dict,
    prefix: str,
):
    """
    Retrieve a metric whose key contains a dynamic k value.

    Example:
        prefix="initial_recall@"

    may match:
        initial_recall@8
    """

    for key, value in metrics.items():
        if key.startswith(prefix):
            return value

    return None


def save_evaluation_excel(
    *,
    summary: dict,
    detailed_results: list[dict],
    output_path: Path,
) -> None:
    """
    Save evaluation results to an Excel workbook.

    Sheets:
        Summary
        Detailed Results
        Errors
    """

    workbook = Workbook()

    # =========================================================
    # Summary sheet
    # =========================================================

    summary_sheet = workbook.active
    summary_sheet.title = "Summary"

    summary_sheet.append(
        [
            "Metric",
            "Mean Score",
        ]
    )

    for metric, value in summary.items():
        summary_sheet.append(
            [
                metric,
                value,
            ]
        )

    _style_header(
        summary_sheet
    )

    summary_sheet.freeze_panes = "A2"

    _autosize_columns(
        summary_sheet,
        max_width=40,
    )

    # =========================================================
    # Detailed results sheet
    # =========================================================

    details_sheet = workbook.create_sheet(
        "Detailed Results"
    )

    headers = [
        "id",
        "category",
        "question",
        "reference_answer",
        "generated_answer",

        "gold_source_ids",
        "retrieved_source_ids",

        "initial_precision",
        "initial_recall",
        "initial_hit_rate",
        "initial_mrr",
        "initial_ndcg",

        "final_precision",
        "final_recall",
        "final_hit_rate",
        "final_mrr",
        "final_ndcg",

        "answer_token_f1",

        "answer_correctness",
        "correctness_reason",

        "faithfulness",
        "faithfulness_reason",

        "answer_relevance",
        "relevance_reason",

        "latency_seconds",

        "tool_count",
        "tool_success_rate",
        "tool_f1",

        "tool_calls",

        "error",
    ]

    details_sheet.append(
        headers
    )

    for item in detailed_results:
        metrics = item.get(
            "metrics",
            {},
        )

        reasons = item.get(
            "judge_reasons",
            {},
        )

        details_sheet.append(
            [
                item.get("id"),
                item.get("category"),
                item.get("question"),
                item.get(
                    "reference_answer"
                ),
                item.get(
                    "generated_answer"
                ),

                _json_string(
                    item.get(
                        "gold_source_ids",
                        [],
                    )
                ),

                _json_string(
                    item.get(
                        "retrieved_source_ids",
                        [],
                    )
                ),

                _get_metric_by_prefix(
                    metrics,
                    "initial_precision@",
                ),

                _get_metric_by_prefix(
                    metrics,
                    "initial_recall@",
                ),

                _get_metric_by_prefix(
                    metrics,
                    "initial_hit_rate@",
                ),

                metrics.get(
                    "initial_mrr"
                ),

                _get_metric_by_prefix(
                    metrics,
                    "initial_ndcg@",
                ),

                _get_metric_by_prefix(
                    metrics,
                    "final_precision@",
                ),

                _get_metric_by_prefix(
                    metrics,
                    "final_recall@",
                ),

                _get_metric_by_prefix(
                    metrics,
                    "final_hit_rate@",
                ),

                metrics.get(
                    "final_mrr"
                ),

                _get_metric_by_prefix(
                    metrics,
                    "final_ndcg@",
                ),

                metrics.get(
                    "answer_token_f1"
                ),

                metrics.get(
                    "answer_correctness"
                ),

                reasons.get(
                    "correctness"
                ),

                metrics.get(
                    "faithfulness"
                ),

                reasons.get(
                    "faithfulness"
                ),

                metrics.get(
                    "answer_relevance"
                ),

                reasons.get(
                    "relevance"
                ),

                metrics.get(
                    "latency_seconds"
                ),

                metrics.get(
                    "tool_count"
                ),

                metrics.get(
                    "tool_success_rate"
                ),

                metrics.get(
                    "tool_f1"
                ),

                _json_string(
                    item.get(
                        "tool_calls",
                        [],
                    )
                ),

                item.get(
                    "error"
                ),
            ]
        )

    _style_header(
        details_sheet
    )

    details_sheet.freeze_panes = "A2"

    details_sheet.auto_filter.ref = (
        details_sheet.dimensions
    )

    for row in details_sheet.iter_rows(
        min_row=2
    ):
        for cell in row:
            cell.alignment = Alignment(
                vertical="top",
                wrap_text=True,
            )

    _autosize_columns(
        details_sheet,
        max_width=45,
    )

    # =========================================================
    # Errors sheet
    # =========================================================

    errors_sheet = workbook.create_sheet(
        "Errors"
    )

    errors_sheet.append(
        [
            "id",
            "category",
            "question",
            "reference_answer",
            "error",
        ]
    )

    for item in detailed_results:
        error = item.get(
            "error"
        )

        if not error:
            continue

        errors_sheet.append(
            [
                item.get("id"),
                item.get("category"),
                item.get("question"),
                item.get(
                    "reference_answer"
                ),
                error,
            ]
        )

    _style_header(
        errors_sheet
    )

    errors_sheet.freeze_panes = "A2"

    for row in errors_sheet.iter_rows(
        min_row=2
    ):
        for cell in row:
            cell.alignment = Alignment(
                vertical="top",
                wrap_text=True,
            )

    _autosize_columns(
        errors_sheet,
        max_width=60,
    )

    # =========================================================
    # Save workbook
    # =========================================================

    output_path.parent.mkdir(
        parents=True,
        exist_ok=True,
    )

    workbook.save(
        output_path
    )


def evaluate_pipeline(
    *,
    dataset_path: str,
    pipeline: str,
    retrieval_mode: str = "hybrid",
    top_k: int = 8,
    limit: int | None = None,
    output_dir: str = "outputs/evaluation",
    model: str | None = None,
    judge_model: str | None = None,
) -> dict:
    """
    Evaluate either the RAG or agent pipeline.

    Metrics include:

    Retrieval:
        - Precision@k
        - Recall@k
        - Hit Rate@k
        - MRR
        - nDCG@k

    Generation:
        - Token F1
        - Correctness
        - Faithfulness
        - Answer relevance

    Efficiency:
        - Latency

    Agent-only:
        - Tool count
        - Tool success rate
        - Tool F1 when expected tools are annotated

    Results are written to a single Excel workbook.
    """

    examples = load_dataset(
        dataset_path
    )

    if limit is not None:
        examples = examples[
            :limit
        ]

    # =========================================================
    # Pipeline adapter
    # =========================================================

    if pipeline == "rag":
        adapter = RAGAdapter(
            mode=retrieval_mode,
            top_k=top_k,
            model=model,
        )

    elif pipeline == "agent":
        adapter = AgentAdapter(
            top_k=top_k,
            model=model,
        )

    else:
        raise ValueError(
            "pipeline must be 'rag' or 'agent'"
        )

    # =========================================================
    # LLM judge
    # =========================================================

    judge = GeminiJudge(
        model=judge_model
    )

    detailed_results = []

    # =========================================================
    # Evaluate examples
    # =========================================================

    for index, example in enumerate(
        examples,
        start=1,
    ):

        print(
            f"[{index}/{len(examples)}] "
            f"{example.question}"
        )

        try:
            # -------------------------------------------------
            # Run pipeline
            # -------------------------------------------------

            result = adapter.run(
                example
            )

            # -------------------------------------------------
            # Initial retrieval metrics
            # -------------------------------------------------

            initial_retrieval = (
                retrieval_metrics(
                    result.initial_retrieved_source_ids,
                    example.gold_source_ids,
                    k=top_k,
                )
            )

            # -------------------------------------------------
            # Final retrieval metrics
            # -------------------------------------------------

            final_retrieval = (
                retrieval_metrics(
                    result.retrieved_source_ids,
                    example.gold_source_ids,
                    k=top_k,
                )
            )

            # -------------------------------------------------
            # LLM-as-a-judge
            # -------------------------------------------------

            judge_scores = (
                judge.evaluate(
                    question=(
                        example.question
                    ),
                    reference_answer=(
                        example.reference_answer
                    ),
                    generated_answer=(
                        result.answer
                    ),
                    contexts=(
                        result.contexts
                    ),
                )
            )

            # -------------------------------------------------
            # Combined metrics
            # -------------------------------------------------

            metrics = {
                **{
                    f"initial_{key}": value
                    for key, value
                    in initial_retrieval.items()
                },

                **{
                    f"final_{key}": value
                    for key, value
                    in final_retrieval.items()
                },

                "answer_token_f1": (
                    token_f1(
                        result.answer,
                        example.reference_answer,
                    )
                ),

                "answer_correctness": float(
                    judge_scores[
                        "correctness"
                    ][
                        "score"
                    ]
                ),

                "faithfulness": float(
                    judge_scores[
                        "faithfulness"
                    ][
                        "score"
                    ]
                ),

                "answer_relevance": float(
                    judge_scores[
                        "relevance"
                    ][
                        "score"
                    ]
                ),

                "latency_seconds": (
                    result.latency_seconds
                ),
            }

            # -------------------------------------------------
            # Agent-specific metrics
            # -------------------------------------------------

            if pipeline == "agent":
                metrics.update(
                    agent_metrics(
                        result.tool_calls,
                        example.expected_tools,
                    )
                )

            # -------------------------------------------------
            # Store detailed result
            # -------------------------------------------------

            detailed_results.append(
                {
                    "id": (
                        example.id
                    ),

                    "category": (
                        example.category
                    ),

                    "question": (
                        example.question
                    ),

                    "reference_answer": (
                        example.reference_answer
                    ),

                    "generated_answer": (
                        result.answer
                    ),

                    "gold_source_ids": (
                        example.gold_source_ids
                    ),

                    "retrieved_source_ids": (
                        result.retrieved_source_ids
                    ),

                    "judge_reasons": {
                        "correctness": (
                            judge_scores[
                                "correctness"
                            ][
                                "reason"
                            ]
                        ),

                        "faithfulness": (
                            judge_scores[
                                "faithfulness"
                            ][
                                "reason"
                            ]
                        ),

                        "relevance": (
                            judge_scores[
                                "relevance"
                            ][
                                "reason"
                            ]
                        ),
                    },

                    "metrics": (
                        metrics
                    ),

                    "tool_calls": (
                        result.tool_calls
                    ),

                    "error": None,
                }
            )

        except Exception as exc:

            detailed_results.append(
                {
                    "id": (
                        example.id
                    ),

                    "category": (
                        example.category
                    ),

                    "question": (
                        example.question
                    ),

                    "reference_answer": (
                        example.reference_answer
                    ),

                    "generated_answer": None,

                    "gold_source_ids": (
                        example.gold_source_ids
                    ),

                    "retrieved_source_ids": [],

                    "judge_reasons": {},

                    "metrics": {},

                    "tool_calls": [],

                    "error": str(exc),
                }
            )

    # =========================================================
    # Aggregate metrics
    # =========================================================

    flat_metrics = [
        item["metrics"]
        for item in detailed_results
        if item.get(
            "metrics"
        )
    ]

    summary = (
        _mean_numeric_metrics(
            flat_metrics
        )
    )

    successful_examples = len(
        flat_metrics
    )

    failed_examples = (
        len(examples)
        - successful_examples
    )

    summary[
        "successful_examples"
    ] = successful_examples

    summary[
        "failed_examples"
    ] = failed_examples

    if examples:
        summary[
            "evaluation_success_rate"
        ] = (
            successful_examples
            / len(examples)
        )

    # =========================================================
    # Output filename
    # =========================================================

    output_directory = Path(
        output_dir
    )

    if pipeline == "rag":
        run_name = (
            f"rag_{retrieval_mode}"
        )
    else:
        run_name = "agent"

    excel_path = (
        output_directory
        / f"{run_name}_evaluation.xlsx"
    )

    # =========================================================
    # Save Excel workbook
    # =========================================================

    save_evaluation_excel(
        summary=summary,
        detailed_results=(
            detailed_results
        ),
        output_path=excel_path,
    )

    return {
        "pipeline": pipeline,
        "retrieval_mode": (
            retrieval_mode
            if pipeline == "rag"
            else None
        ),
        "examples": len(
            examples
        ),
        "successful_examples": (
            successful_examples
        ),
        "failed_examples": (
            failed_examples
        ),
        "summary": summary,
        "excel_file": str(
            excel_path
        ),
    }